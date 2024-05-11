import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By
import os
from dotenv import load_dotenv

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def test_postal_upto_10books():
    driver = webdriver.Chrome()
    load_dotenv()
    driver.get(os.getenv("LINK"))
    driver.maximize_window()
    WebDriverWait(driver, 10).until(
        EC.text_to_be_present_in_element((By.XPATH, "//span[contains(text(),'Agent Login')]"), "Agent Login")
    )
    agent_id = driver.find_element(By.XPATH, "//input[@name ='AuthenticationFG.USER_PRINCIPAL']")
    password = driver.find_element(By.XPATH, "//input[@name ='AuthenticationFG.ACCESS_CODE']")
    log_in = driver.find_element(By.XPATH, "//input[@value ='Log in']")
    agent_id.send_keys(os.getenv("USERID"))
    password.send_keys(os.getenv("PASSWORD"))
    time.sleep(10)
    log_in.click()
    WebDriverWait(driver, 25).until(
        EC.text_to_be_present_in_element((By.XPATH, "//a[@id='Accounts']"), "Accounts")
    )
    accounts = driver.find_element(By.XPATH, "//a[@id='Accounts']")
    accounts.click()
    WebDriverWait(driver, 25).until(
        EC.text_to_be_present_in_element((By.XPATH, "//a[@id='Agent Enquire & Update Screen']"), "Agent Enquire & Update Screen")
    )
    books_screen = driver.find_element(By.XPATH, "//a[@id='Agent Enquire & Update Screen']")
    books_screen.click()
    WebDriverWait(driver, 25).until(
        EC.text_to_be_present_in_element((By.XPATH, "//span[contains(text(),'Select Mode:')]"),
                                         "Select Mode:")
    )
    cash = driver.find_element(By.XPATH,"//input[@id = 'CustomAgentRDAccountFG.PAY_MODE_SELECTED_FOR_TRN' and @value ='C']")
    cash.click()
    account_id = driver.find_element(By.NAME, "CustomAgentRDAccountFG.ACCOUNT_NUMBER_FOR_SEARCH")
    file_path = "/Users/Ganesh/Postal-Automation/Postal/Data/Account_Numbers.xlsx"
    workbook = load_workbook(file_path)
    sheet = workbook['Sheet2']
    account_no = sheet.cell(row=1, column=1).value
    account_id.send_keys(account_no)
    fetch = driver.find_element(By.XPATH,"//input[@name='Action.FETCH_INPUT_ACCOUNT']")
    fetch.click()
    WebDriverWait(driver, 25).until(
        EC.text_to_be_present_in_element((By.XPATH, "//span[contains(text(),'Select Mode:')]"),
                                         "Select Mode:")
    )
    select_book = driver.find_elements(By.XPATH,"//input[@type='checkbox']")
    for checkbox in select_book:
        checkbox.click()
    #time.sleep(40)
    save = driver.find_element(By.XPATH,"//input[@name='Action.SAVE_ACCOUNTS']")
    save.click()
    # WebDriverWait(driver, 25).until(
    #     EC.text_to_be_present_in_element((By.XPATH, "//span[contains(text(),'Select Mode:')]"),
    #                                      "Select Mode:")
    # )
    # pay_all = driver.find_element(By.XPATH,"//input[@value='Pay All Saved Installments']")
    # pay_all.click()
    # WebDriverWait(driver, 25).until(
    #     EC.text_to_be_present_in_element((By.XPATH, "//span[contains(text(),'Selected Recurring Deposit Account List')]"),
    #                                      "Selected Recurring Deposit Account List")
    # )
    time.sleep(5)
    driver.save_screenshot("screenshot.png")

    time.sleep(4 * 60 * 60)  # 4 hours



